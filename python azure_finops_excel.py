#!/usr/bin/env python3
"""
Azure FinOps - Analyse des coûts par Resource Group (30 derniers jours)
Génère un fichier Excel multi-souscriptions à la place d'un push Notion.

Authentification :
  - Sur Azure (VM, Function, Container…) : Managed Identity automatique
  - En local : az login  (Azure CLI)

Prérequis:
  pip install azure-identity azure-mgmt-costmanagement requests openpyxl python-dotenv

Usage:
  python azure_finops_excel.py                              # toutes les souscriptions
  python azure_finops_excel.py --subscription-id <ID>,...   # filtre sur une ou plusieurs
  python azure_finops_excel.py --output mon_rapport.xlsx
"""

import argparse
import os
import random
import sys
import time
from datetime import datetime, timedelta, timezone

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from azure.identity import DefaultAzureCredential
    from azure.mgmt.costmanagement import CostManagementClient
except ImportError:
    print("Packages manquants. Installez-les avec:")
    print("  pip install azure-identity azure-mgmt-costmanagement requests openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Package openpyxl manquant. Installez-le avec:  pip install openpyxl")
    sys.exit(1)


# --- Styles Excel ---
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
DATA_FONT = Font(name="Arial", size=10)
CURRENCY_FMT = '#,##0.00 €'
PCT_FMT = '0.0%'
RED_FONT = Font(name="Arial", size=10, color="C00000")
GREEN_FONT = Font(name="Arial", size=10, color="007B2B")
BOLD_FONT = Font(name="Arial", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def get_all_subscriptions(credential):
    """Retourne la liste de toutes les souscriptions accessibles (id, nom)."""
    import requests
    token = credential.get_token("https://management.azure.com/.default").token
    resp = requests.get(
        "https://management.azure.com/subscriptions?api-version=2020-01-01",
        headers={"Authorization": f"Bearer {token}"}
    )
    resp.raise_for_status()
    subs = resp.json().get("value", [])
    if not subs:
        print("Erreur: aucune souscription accessible avec ce credential.")
        sys.exit(1)
    return [(s["subscriptionId"], s["displayName"]) for s in subs]


# Headers Azure qui indiquent combien de temps attendre avant de réessayer
_RETRY_HEADERS = (
    "x-ms-ratelimit-microsoft.costmanagement-entity-retry-after",
    "x-ms-ratelimit-microsoft.costmanagement-tenant-retry-after",
    "x-ms-ratelimit-microsoft.consumption-tenant-retry-after",
    "Retry-After",
)


def _extract_retry_after(exc):
    """Extrait la valeur Retry-After (en secondes) depuis une réponse Azure, ou None."""
    resp = getattr(exc, "response", None)
    if resp is None:
        return None
    headers = getattr(resp, "headers", None) or {}
    for h in _RETRY_HEADERS:
        val = headers.get(h) or headers.get(h.lower())
        if val is not None:
            try:
                return max(1, int(float(val)))
            except (TypeError, ValueError):
                continue
    return None


def query_costs(client, scope, start_date, end_date, grouping_dimensions, max_attempts=6):
    """Interroge l'API Cost Management avec retry robuste (429/503 + Retry-After)."""
    from azure.core.exceptions import HttpResponseError
    from azure.mgmt.costmanagement.models import (
        QueryDefinition, QueryTimePeriod, QueryDataset,
        QueryAggregation, QueryGrouping,
    )

    query = QueryDefinition(
        type="ActualCost",
        timeframe="Custom",
        time_period=QueryTimePeriod(from_property=start_date, to=end_date),
        dataset=QueryDataset(
            granularity="Daily",
            aggregation={"totalCost": QueryAggregation(name="Cost", function="Sum")},
            grouping=[QueryGrouping(type="Dimension", name=dim) for dim in grouping_dimensions]
        )
    )

    for attempt in range(max_attempts):
        try:
            response = client.query.usage(scope=scope, parameters=query, timeout=120)
            columns = [col.name for col in response.columns] if response.columns else []
            results = list(response.rows) if response.rows else []
            return columns, results
        except HttpResponseError as e:
            status = getattr(e, "status_code", None)
            throttled = status in (429, 503) or "429" in str(e) or "503" in str(e)
            if not throttled or attempt == max_attempts - 1:
                raise
            wait = _extract_retry_after(e)
            if wait is None:
                wait = min(120, 5 * (2 ** attempt)) + random.uniform(0, 3)
            print(f"  Rate limit (status={status}), attente {int(wait)}s... ({attempt + 1}/{max_attempts})")
            time.sleep(wait)
        except Exception as e:
            if ("429" in str(e) or "503" in str(e)) and attempt < max_attempts - 1:
                wait = min(120, 5 * (2 ** attempt)) + random.uniform(0, 3)
                print(f"  Rate limit, attente {int(wait)}s... ({attempt + 1}/{max_attempts})")
                time.sleep(wait)
            else:
                raise


def build_resource_group_data(columns, rows):
    """Structure les données par Resource Group et par jour."""
    data = {}
    daily_totals = {}

    for row in rows:
        cost = float(row[0])
        date_val = row[1]
        rg_name = row[2] if len(row) > 2 else "N/A"

        if isinstance(date_val, (int, float)):
            date_str = str(int(date_val))
            date_key = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        else:
            date_key = str(date_val)[:10]

        if rg_name not in data:
            data[rg_name] = {}
        data[rg_name][date_key] = data[rg_name].get(date_key, 0) + cost
        daily_totals[date_key] = daily_totals.get(date_key, 0) + cost

    return data, daily_totals


# --- Helpers Excel ---
def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_cell(cell, kind="text"):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if kind == "currency":
        cell.number_format = CURRENCY_FMT
        cell.alignment = Alignment(horizontal="right")
    elif kind == "pct":
        cell.number_format = PCT_FMT
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.alignment = Alignment(horizontal="left")


def color_evolution(cell, value, threshold=0.05):
    if value > threshold:
        cell.font = RED_FONT
        cell.fill = PatternFill("solid", fgColor="FDE8E8")
    elif value < -threshold:
        cell.font = GREEN_FONT
        cell.fill = PatternFill("solid", fgColor="E8FDE8")


def auto_column_width(ws, min_width=12, max_width=35):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# --- Feuilles Excel ---
def create_global_summary(wb, subscriptions_data, dates, threshold):
    """Feuille Vue Globale : agrégat toutes souscriptions + KPIs."""
    ws = wb.active
    ws.title = "Vue Globale"

    mid = len(dates) // 2
    first_half = set(dates[:mid])
    second_half = set(dates[mid:])

    all_rg = {}
    for _, rg_data in subscriptions_data:
        for rg, daily in rg_data.items():
            all_rg.setdefault(rg, {})
            for d, v in daily.items():
                all_rg[rg][d] = all_rg[rg].get(d, 0) + v

    total_previous = sum(sum(v for k, v in d.items() if k in first_half) for d in all_rg.values())
    total_recent = sum(sum(v for k, v in d.items() if k in second_half) for d in all_rg.values())
    global_total = total_previous + total_recent
    global_evo = (total_recent - total_previous) / total_previous if total_previous > 0 else 0.0

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = "Rapport FinOps Azure — Vue Globale"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Période : {dates[0]} → {dates[-1]}  |  {len(subscriptions_data)} souscription(s)"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    kpis = [
        ("Coût Total (30 jours)", global_total, "currency"),
        ("15 jours précédents", total_previous, "currency"),
        ("15 jours récents", total_recent, "currency"),
        ("Évolution globale", global_evo, "pct"),
    ]
    for i, (label, value, kind) in enumerate(kpis):
        row = 4
        col = 1 + i * 2
        ws.cell(row=row, column=col, value=label).font = Font(name="Arial", bold=True, size=11)
        vcell = ws.cell(row=row + 1, column=col, value=value)
        vcell.font = Font(name="Arial", bold=True, size=14, color="2F5496")
        if kind == "currency":
            vcell.number_format = CURRENCY_FMT
        elif kind == "pct":
            vcell.number_format = PCT_FMT
            if value > threshold:
                vcell.font = Font(name="Arial", bold=True, size=14, color="C00000")
            elif value < -threshold:
                vcell.font = Font(name="Arial", bold=True, size=14, color="007B2B")

    # Totaux par souscription
    start = 8
    ws.cell(row=start, column=1, value="Coût par Souscription").font = Font(name="Arial", bold=True, size=12, color="2F5496")
    headers = ["Souscription", "Coût Total (€)", "15j préc. (€)", "15j récents (€)", "Évolution (%)", "Nb RG"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start + 1, column=col, value=h)
    style_header_row(ws, start + 1, len(headers))

    row = start + 2
    sub_totals = []
    for sub_name, rg_data in subscriptions_data:
        sub_total = sum(sum(d.values()) for d in rg_data.values())
        sub_prev = sum(sum(v for k, v in d.items() if k in first_half) for d in rg_data.values())
        sub_rec = sum(sum(v for k, v in d.items() if k in second_half) for d in rg_data.values())
        sub_evo = (sub_rec - sub_prev) / sub_prev if sub_prev > 0 else (1.0 if sub_rec > 0 else 0.0)
        sub_totals.append((sub_name, sub_total, sub_prev, sub_rec, sub_evo, len(rg_data)))

    sub_totals.sort(key=lambda x: x[1], reverse=True)
    for sub_name, sub_total, sub_prev, sub_rec, sub_evo, nb_rg in sub_totals:
        ws.cell(row=row, column=1, value=sub_name); style_cell(ws.cell(row=row, column=1))
        ws.cell(row=row, column=2, value=sub_total); style_cell(ws.cell(row=row, column=2), "currency")
        ws.cell(row=row, column=3, value=sub_prev); style_cell(ws.cell(row=row, column=3), "currency")
        ws.cell(row=row, column=4, value=sub_rec); style_cell(ws.cell(row=row, column=4), "currency")
        evo_cell = ws.cell(row=row, column=5, value=sub_evo); style_cell(evo_cell, "pct")
        color_evolution(evo_cell, sub_evo, threshold)
        ws.cell(row=row, column=6, value=nb_rg); style_cell(ws.cell(row=row, column=6))
        row += 1

    # Ligne totaux
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=SUM(B{start+2}:B{row-1})")
    ws.cell(row=row, column=3, value=f"=SUM(C{start+2}:C{row-1})")
    ws.cell(row=row, column=4, value=f"=SUM(D{start+2}:D{row-1})")
    ws.cell(row=row, column=5, value=f"=IF(C{row}>0,(D{row}-C{row})/C{row},0)")
    ws.cell(row=row, column=6, value=f"=SUM(F{start+2}:F{row-1})")
    for col in range(2, 7):
        c = ws.cell(row=row, column=col)
        c.font = BOLD_FONT
        c.border = THIN_BORDER
        if col in (2, 3, 4):
            c.number_format = CURRENCY_FMT
        elif col == 5:
            c.number_format = PCT_FMT

    # Graphique Top 10 RG globaux
    chart_start = row + 3
    ws.cell(row=chart_start, column=1, value="Top 10 Resource Groups (toutes souscriptions)").font = Font(name="Arial", bold=True, size=12, color="2F5496")

    top_headers = ["Resource Group", "Coût Total (€)"]
    for col, h in enumerate(top_headers, 1):
        ws.cell(row=chart_start + 1, column=col, value=h)
    style_header_row(ws, chart_start + 1, len(top_headers))

    top_rg = sorted(all_rg.items(), key=lambda x: sum(x[1].values()), reverse=True)[:10]
    for i, (rg_name, daily) in enumerate(top_rg):
        r = chart_start + 2 + i
        ws.cell(row=r, column=1, value=rg_name); style_cell(ws.cell(row=r, column=1))
        ws.cell(row=r, column=2, value=sum(daily.values())); style_cell(ws.cell(row=r, column=2), "currency")

    if top_rg:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top 10 Resource Groups — Coût Total (€)"
        chart.style = 10
        chart.width = 25
        chart.height = 12
        data_ref = Reference(ws, min_col=2, min_row=chart_start + 1, max_row=chart_start + 1 + len(top_rg))
        cats_ref = Reference(ws, min_col=1, min_row=chart_start + 2, max_row=chart_start + 1 + len(top_rg))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "2F5496"
        ws.add_chart(chart, f"D{chart_start}")

    auto_column_width(ws)
    ws.sheet_properties.tabColor = "2F5496"


def create_alerts_sheet(wb, subscriptions_data, dates, threshold):
    """Feuille Alertes : tous les RG en hausse supérieure au seuil."""
    ws = wb.create_sheet("⚠ Alertes Hausses")

    ws["A1"] = f"Alertes — Resource Groups en hausse de plus de {int(threshold * 100)}%"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="C00000")
    ws.merge_cells("A1:G1")

    headers = ["Souscription", "Resource Group", "Coût Total (€)",
               "15j préc. (€)", "15j récents (€)", "Hausse (€)", "Évolution (%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    mid = len(dates) // 2
    first_half = set(dates[:mid])
    second_half = set(dates[mid:])

    alerts = []
    for sub_name, rg_data in subscriptions_data:
        for rg_name, daily in rg_data.items():
            previous = sum(v for k, v in daily.items() if k in first_half)
            recent = sum(v for k, v in daily.items() if k in second_half)
            evo = (recent - previous) / previous if previous > 0 else (1.0 if recent > 0 else 0.0)
            if evo > threshold:
                alerts.append((sub_name, rg_name, sum(daily.values()), previous, recent, recent - previous, evo))

    alerts.sort(key=lambda x: x[6], reverse=True)

    if not alerts:
        ws["A4"] = f"Aucun Resource Group en hausse de plus de {int(threshold * 100)}%."
        ws["A4"].font = Font(name="Arial", italic=True, color="007B2B")
    else:
        for i, (sub_name, rg_name, total, prev, rec, hausse, evo) in enumerate(alerts):
            row = 4 + i
            ws.cell(row=row, column=1, value=sub_name); style_cell(ws.cell(row=row, column=1))
            ws.cell(row=row, column=2, value=rg_name); style_cell(ws.cell(row=row, column=2))
            ws.cell(row=row, column=3, value=total); style_cell(ws.cell(row=row, column=3), "currency")
            ws.cell(row=row, column=4, value=prev); style_cell(ws.cell(row=row, column=4), "currency")
            ws.cell(row=row, column=5, value=rec); style_cell(ws.cell(row=row, column=5), "currency")
            hausse_cell = ws.cell(row=row, column=6, value=hausse); style_cell(hausse_cell, "currency")
            hausse_cell.font = RED_FONT
            evo_cell = ws.cell(row=row, column=7, value=evo); style_cell(evo_cell, "pct")
            evo_cell.font = RED_FONT
            evo_cell.fill = PatternFill("solid", fgColor="FDE8E8")

    auto_column_width(ws)
    ws.sheet_properties.tabColor = "C00000"


def create_subscription_sheet(wb, sub_name, rg_data, dates, threshold):
    """Feuille détaillée par souscription : RG + J-3/J-2 + tendance 15j."""
    # Nom d'onglet Excel : max 31 chars, pas de caractères spéciaux
    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in sub_name)[:31]
    ws = wb.create_sheet(safe_name or "Subscription")

    mid = len(dates) // 2
    first_half = set(dates[:mid])
    second_half = set(dates[mid:])
    day_j2 = dates[-2] if len(dates) >= 2 else None
    day_j3 = dates[-3] if len(dates) >= 3 else None

    ws["A1"] = f"🔷 {sub_name}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws.merge_cells("A1:I1")

    sub_total = sum(sum(d.values()) for d in rg_data.values())
    ws["A2"] = f"Coût total : {sub_total:,.2f} €  |  {len(rg_data)} resource groups"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws.merge_cells("A2:I2")

    headers = [
        "Resource Group", "Coût Total (€)", "Moy/Jour (€)",
        "15j préc. (€)", "15j récents (€)", "Évolution (%)",
        f"J-3 ({day_j3})" if day_j3 else "J-3",
        f"J-2 ({day_j2})" if day_j2 else "J-2",
        "Évo. J-3→J-2",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    sorted_rgs = sorted(rg_data.items(), key=lambda x: sum(x[1].values()), reverse=True)
    for i, (rg_name, daily) in enumerate(sorted_rgs):
        row = 5 + i
        total = sum(daily.values())
        avg = total / len(dates)
        previous = sum(v for k, v in daily.items() if k in first_half)
        recent = sum(v for k, v in daily.items() if k in second_half)
        evo = (recent - previous) / previous if previous > 0 else (1.0 if recent > 0 else 0.0)
        cost_j3 = daily.get(day_j3, 0) if day_j3 else 0
        cost_j2 = daily.get(day_j2, 0) if day_j2 else 0
        evo_daily = (cost_j2 - cost_j3) / cost_j3 if cost_j3 > 0 else (1.0 if cost_j2 > 0 else 0.0)

        ws.cell(row=row, column=1, value=rg_name); style_cell(ws.cell(row=row, column=1))
        ws.cell(row=row, column=2, value=total); style_cell(ws.cell(row=row, column=2), "currency")
        ws.cell(row=row, column=3, value=avg); style_cell(ws.cell(row=row, column=3), "currency")
        ws.cell(row=row, column=4, value=previous); style_cell(ws.cell(row=row, column=4), "currency")
        ws.cell(row=row, column=5, value=recent); style_cell(ws.cell(row=row, column=5), "currency")
        evo_cell = ws.cell(row=row, column=6, value=evo); style_cell(evo_cell, "pct")
        color_evolution(evo_cell, evo, threshold)

        ws.cell(row=row, column=7, value=cost_j3); style_cell(ws.cell(row=row, column=7), "currency")
        ws.cell(row=row, column=8, value=cost_j2); style_cell(ws.cell(row=row, column=8), "currency")
        evo_d_cell = ws.cell(row=row, column=9, value=evo_daily); style_cell(evo_d_cell, "pct")
        color_evolution(evo_d_cell, evo_daily, 0.05)

    # Ligne totaux
    total_row = 5 + len(sorted_rgs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=AVERAGE(C5:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row - 1})")
    ws.cell(row=total_row, column=6, value=f"=IF(D{total_row}>0,(E{total_row}-D{total_row})/D{total_row},0)")
    ws.cell(row=total_row, column=7, value=f"=SUM(G5:G{total_row - 1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H5:H{total_row - 1})")
    ws.cell(row=total_row, column=9, value=f"=IF(G{total_row}>0,(H{total_row}-G{total_row})/G{total_row},0)")
    for col in range(2, 10):
        c = ws.cell(row=total_row, column=col)
        c.font = BOLD_FONT
        c.border = THIN_BORDER
        if col in (2, 3, 4, 5, 7, 8):
            c.number_format = CURRENCY_FMT
        elif col in (6, 9):
            c.number_format = PCT_FMT

    auto_column_width(ws)
    ws.sheet_properties.tabColor = "27AE60"


def create_daily_evolution_sheet(wb, subscriptions_data, dates):
    """Feuille Évolution Quotidienne : totaux journaliers + courbe."""
    ws = wb.create_sheet("Évolution Quotidienne")

    ws["A1"] = "Évolution quotidienne des coûts (toutes souscriptions)"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws.merge_cells("A1:D1")

    # Agréger par date
    daily_all = {d: 0 for d in dates}
    for _, rg_data in subscriptions_data:
        for daily in rg_data.values():
            for d, v in daily.items():
                if d in daily_all:
                    daily_all[d] += v

    headers = ["Date", "Coût Quotidien (€)", "Variation vs Veille (€)", "Variation (%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    for i, date in enumerate(dates):
        row = 4 + i
        cost = daily_all[date]
        prev_cost = daily_all[dates[i - 1]] if i > 0 else 0

        ws.cell(row=row, column=1, value=date); style_cell(ws.cell(row=row, column=1))
        ws.cell(row=row, column=2, value=cost); style_cell(ws.cell(row=row, column=2), "currency")

        if i > 0:
            variation = cost - prev_cost
            var_cell = ws.cell(row=row, column=3, value=variation); style_cell(var_cell, "currency")
            if variation > 0:
                var_cell.font = RED_FONT
            elif variation < 0:
                var_cell.font = GREEN_FONT
            pct = variation / prev_cost if prev_cost > 0 else 0
            pct_cell = ws.cell(row=row, column=4, value=pct); style_cell(pct_cell, "pct")
            if pct > 0:
                pct_cell.font = RED_FONT
            elif pct < 0:
                pct_cell.font = GREEN_FONT

    last_row = 3 + len(dates)

    chart = LineChart()
    chart.title = "Évolution quotidienne des coûts (€)"
    chart.style = 10
    chart.width = 30
    chart.height = 12
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = "2F5496"
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, f"F3")

    auto_column_width(ws)
    ws.sheet_properties.tabColor = "E67E22"


def export_to_excel(subscriptions_data, dates, output_path, threshold=0.05):
    """Génère le fichier Excel multi-onglets."""
    wb = Workbook()
    create_global_summary(wb, subscriptions_data, dates, threshold)
    create_daily_evolution_sheet(wb, subscriptions_data, dates)
    create_alerts_sheet(wb, subscriptions_data, dates, threshold)
    for sub_name, rg_data in subscriptions_data:
        create_subscription_sheet(wb, sub_name, rg_data, dates, threshold)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Azure FinOps — Export Excel multi-souscriptions")
    parser.add_argument("--subscription-id", default=os.getenv("AZURE_SUBSCRIPTION_ID"),
                        help="ID(s) de souscription séparés par des virgules (défaut : toutes)")
    parser.add_argument("--output", default=None,
                        help="Chemin du fichier Excel de sortie (défaut : azure_finops_YYYYMMDD.xlsx)")
    parser.add_argument("--threshold", type=float, default=0.05,
                        help="Seuil d'alerte d'évolution (défaut : 0.05 = 5%%)")
    args = parser.parse_args()

    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    output_path = args.output or f"azure_finops_{today}.xlsx"

    print("Authentification (Managed Identity / Azure CLI)...")
    credential = DefaultAzureCredential()

    all_subs = get_all_subscriptions(credential)
    if args.subscription_id:
        filter_ids = {s.strip() for s in args.subscription_id.split(",")}
        all_subs = [(sid, name) for sid, name in all_subs if sid in filter_ids]
        if not all_subs:
            print("Aucune souscription correspondante trouvée.")
            sys.exit(1)

    print(f"{len(all_subs)} souscription(s) à traiter :")
    for sid, name in all_subs:
        print(f"  {sid}  {name}")

    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=30)
    print(f"\nPériode : {start_date.strftime('%Y-%m-%d')} → {end_date.strftime('%Y-%m-%d')}")
    dates = [(start_date + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(31)]

    subscriptions_data = []
    client = CostManagementClient(credential)

    # Azure Cost Management : ~15 requêtes/min par scope. On espace pour rester sous la limite.
    SUB_DELAY_SECONDS = float(os.getenv("SUB_DELAY_SECONDS", "4"))

    for idx, (sub_id, sub_name) in enumerate(all_subs):
        if idx > 0 and SUB_DELAY_SECONDS > 0:
            time.sleep(SUB_DELAY_SECONDS)
        print(f"\n[{sub_name}] Récupération des coûts...")
        try:
            rg_columns, rg_rows = query_costs(
                client, f"/subscriptions/{sub_id}", start_date, end_date, ["ResourceGroup"]
            )
            rg_data, _ = build_resource_group_data(rg_columns, rg_rows)
            print(f"  → {len(rg_data)} resource groups")
            if rg_data:
                subscriptions_data.append((sub_name, rg_data))
        except Exception as e:
            err = str(e)
            if "BillingAccount" in err or "IndirectCostDisabled" in err or "Unauthorized" in err or "404" in err:
                print(f"  → Ignorée (Cost Management non disponible sur cette souscription)")
            elif "timeout" in err.lower():
                print(f"  → Ignorée (timeout)")
            else:
                print(f"  → Ignorée : {err[:120]}")

    if not subscriptions_data:
        print("Aucune donnée de coûts trouvée.")
        sys.exit(1)

    # Réduire les dates aux jours réellement présents
    all_dates_in_data = set()
    for _, rg_data in subscriptions_data:
        for daily in rg_data.values():
            all_dates_in_data.update(daily.keys())
    dates = sorted([d for d in dates if d in all_dates_in_data])

    print(f"\nGénération du fichier Excel : {output_path}")
    export_to_excel(subscriptions_data, dates, output_path, args.threshold)
    print("\n✅ Rapport généré avec succès.")
    print(f"   • Vue Globale (KPIs + Top 10 RG)")
    print(f"   • Évolution Quotidienne (courbe)")
    print(f"   • ⚠ Alertes Hausses (> {int(args.threshold * 100)}%)")
    print(f"   • 1 onglet par souscription (détail RG + J-3/J-2)")


if __name__ == "__main__":
    main()
